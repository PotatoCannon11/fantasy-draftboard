#!/usr/bin/env python
"""Prove fantasydraft/live.py reproduces the workbook's live engine.

The xlsx stays as the draft-day backup, so the two surfaces must not drift. This
recalculates the real workbook through LibreOffice and asserts the Python engine
lands on the same numbers from the same draft state - plus the ordinary unit
tests that were impossible while the logic lived in worksheet formulas.

    FDS_SOFFICE=/opt/libreoffice26.2/program/soffice .venv/bin/python test_live.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pandas as pd

# src must win over the root shims, which only re-export these names
sys.path.insert(0, str(Path(__file__).resolve().parent / "fantasydraft"))

from common import DATA_PROC, FANTASY_POS, latest_board, load_config  # noqa: E402
from live import DraftState, LiveBoard  # noqa: E402
from verify_xlsx import COL, find_soffice, sheet_to_rows  # noqa: E402

FAILS: list[str] = []
PASSES = 0


def check(name: str, cond: bool, detail: str = "") -> None:
    global PASSES
    if cond:
        PASSES += 1
        print(f"  [pass] {name}")
    else:
        FAILS.append(f"{name}: {detail}")
        print(f"  [FAIL] {name}  {detail}")


def close(a: float, b: float, tol: float = 0.15) -> bool:
    return abs(a - b) <= tol


def sheet_live(soffice: str, xlsx: Path, work: Path) -> dict:
    """Recalculate the workbook and read its Pick Assistant live panel."""
    import openpyxl
    names = openpyxl.load_workbook(xlsx, read_only=True).sheetnames
    idx = {n: i + 1 for i, n in enumerate(names)}
    rows = sheet_to_rows(soffice, xlsx, idx["Pick Assistant"], work)
    out = {"positions": {}}
    for r in rows:
        if len(r) < 11:
            continue
        head = r[9].strip()
        if head == "Picks made":
            out["picks_made"] = r[10].strip()
        elif head == "Your NEXT turn":
            out["next_turn"] = r[10].strip()
        elif head.startswith("TAKE:") or head.startswith("ROSTER FULL"):
            out["headline"] = head
        elif head in FANTASY_POS and len(r) > 14:
            out["positions"][head] = {
                "name": r[10].strip(), "vbd": r[11].strip(),
                "vona": r[12].strip(), "verdict": r[13].strip(),
                "yours": r[14].strip()}
    return out


def mark(xlsx: Path, dst: Path, drafted: list[str], mine: list[str]) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Live Draft Tracker"]
    want = {n: ("mine" if n in mine else "drafted") for n in drafted + mine}
    for r in range(5, ws.max_row + 1):
        nm = ws.cell(r, COL["player_name"]).value
        if nm and str(nm).strip() in want:
            ws.cell(r, COL[want[str(nm).strip()]]).value = "X"
    wb.save(dst)


def main() -> int:
    cfg = load_config()
    board = pd.read_parquet(DATA_PROC / "draft_board.parquet")
    xlsx = latest_board()

    # The workbook is built from a filtered board; the engine must see exactly
    # the same universe or the queues differ and nothing will line up.
    ocfg = cfg["output"]
    b = board.sort_values("vbd_score", ascending=False).reset_index(drop=True)
    b["overall_rank"] = range(1, len(b) + 1)
    keep = b["n_sources"].fillna(0) >= 2
    keep |= b["overall_rank"] <= ocfg.get("single_source_rank_grace", 60)
    if ocfg.get("single_source_keep_if_adp", True):
        keep |= b[f"adp_{ocfg['adp_teams']}"].notna()
    b = b[keep].head(ocfg["board_depth"]).reset_index(drop=True)

    lb = LiveBoard(b, cfg)
    soffice = find_soffice()

    print("=== unit tests (no spreadsheet needed) ===")
    st = DraftState()
    check("empty draft: 0 picks made", lb.picks_made(st) == 0)
    check("empty draft: on the clock is pick 1", lb.on_clock(st) == 1)
    this, nxt = lb.turns(st)
    check("first turn is your slot", this == cfg["simulation"]["draft_slot"],
          f"got {this}")
    check("next turn is the snake mirror",
          nxt == 2 * lb.n_teams - lb.slot + 1, f"got {nxt}")

    rec = lb.recommend(st)
    check("empty board recommends someone", rec["name"] is not None)
    check("survival probability is a probability",
          all(0.0 <= d["survival"] <= 1.0 for d in rec["positions"].values()))

    # Undo must be exact.
    top = int(b.index[0])
    st2 = DraftState()
    st2.take(top, mine=True)
    check("taking a player removes him", not lb.available(st2)[top])
    st2.undo()
    check("undo restores him", lb.available(st2)[top])
    check("undo restores the pick count", lb.picks_made(st2) == 0)

    # Roster awareness: fill QB and it must stop being recommended.
    cap = int(ocfg["useful_max"]["QB"])
    st3 = DraftState()
    qbs = [int(i) for i in b.index[b.position == "QB"][:cap]]
    for i in qbs:
        st3.take(i, mine=True)
    v3 = lb.vona(st3)
    check("a filled position reports full", v3["QB"]["full"])
    check("a filled position counts your picks", v3["QB"]["yours"] == cap)
    check("a filled position drops out of the headline",
          lb.recommend(st3)["pos"] != "QB",
          f"still recommends {lb.recommend(st3)['pos']}")

    # Run detector - genuinely new, the sheet could not see pick order.
    st4 = DraftState()
    rbs = [int(i) for i in b.index[b.position == "RB"][:4]]
    wrs = [int(i) for i in b.index[b.position == "WR"][:2]]
    for i in rbs + wrs:
        st4.take(i)
    runs = lb.runs(st4, window=6)
    check("run detector counts the window", runs["RB"] == 4 and runs["WR"] == 2,
          f"got {runs}")
    runs2 = lb.runs(st4, window=2)
    check("run detector respects a shorter window",
          runs2["WR"] == 2 and runs2["RB"] == 0, f"got {runs2}")

    sc = lb.scarcity(DraftState())
    check("scarcity: best tier on a full board is 1",
          all(d["best_tier"] == 1 for d in sc.values()),
          f"{ {p: d['best_tier'] for p, d in sc.items()} }")

    # The unit tests above stand on their own. Parity additionally needs a
    # built workbook AND LibreOffice; where either is missing (CI, a fresh
    # checkout) say so loudly and skip rather than failing the suite for the
    # absence of an optional dependency.
    if xlsx is None or not soffice:
        why = "no workbook in output/" if xlsx is None else "LibreOffice not found"
        print(f"\n[skip] workbook parity - {why}. "
              "Build one with `python -m fantasydraft.run_pipeline` and set "
              "FDS_SOFFICE to check the two surfaces still agree.")
        return _report()

    print("\n=== parity with the workbook's recalculated engine ===")
    scenarios = [
        ("empty board", [], []),
        ("15 gone, 2 QBs mine",
         [str(n) for n in b.nsmallest(15, f"adp_{ocfg['adp_teams']}")
          .player_name],
         [str(n) for n in b[b.position == "QB"].head(cap).player_name]),
    ]
    with tempfile.TemporaryDirectory() as td:
        work = Path(td)
        for label, drafted, mine in scenarios:
            drafted = [d for d in drafted if d not in mine]
            tmp = work / "scratch.xlsx"
            mark(xlsx, tmp, drafted, mine)
            sheet = sheet_live(soffice, tmp, work)

            st = DraftState()
            name_to_row = {str(n): i for i, n in enumerate(b.player_name)}
            for n in drafted:
                if n in name_to_row:
                    st.take(name_to_row[n])
            for n in mine:
                if n in name_to_row:
                    st.take(name_to_row[n], mine=True)

            print(f"  -- {label} --")
            print(f"     sheet : {sheet.get('headline', '')}")
            py = lb.recommend(st)
            print(f"     python: {py['headline']}")
            check(f"{label}: picks made agrees",
                  sheet.get("picks_made") == str(lb.picks_made(st)),
                  f"sheet={sheet.get('picks_made')} py={lb.picks_made(st)}")
            check(f"{label}: same headline call",
                  sheet.get("headline", "").split("(")[0].strip()
                  == py["headline"].split("(")[0].strip(),
                  f"sheet={sheet.get('headline')!r} py={py['headline']!r}")
            for pos, d in sheet["positions"].items():
                if not d["vona"]:
                    continue
                try:
                    sheet_vona = float(d["vona"])
                except ValueError:
                    continue
                check(f"{label}: {pos} VONA matches",
                      close(sheet_vona, py["positions"][pos]["vona"]),
                      f"sheet={sheet_vona} py={py['positions'][pos]['vona']:.1f}")
    return _report()


def _report() -> int:
    print(f"\n{PASSES} passed, {len(FAILS)} failed")
    for f in FAILS:
        print(f"  - {f}")
    return 1 if FAILS else 0


if __name__ == "__main__":
    sys.exit(main())
