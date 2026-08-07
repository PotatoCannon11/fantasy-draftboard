#!/usr/bin/env python
"""Verify the workbook's live formulas by actually recalculating them.

xlsxwriter stores a placeholder value of 0 for every formula it writes, so
reading the file back with openpyxl proves nothing about whether the formulas
work. This drives a real LibreOffice engine over the file with
recalculate-on-load forced on, then asserts against the computed values.

It also catches the `_xlfn.` trap: functions newer than Excel 2007 (MINIFS,
XLOOKUP, IFS, ...) must be written with that prefix in the xlsx format or they
evaluate to #NAME? in every spreadsheet app.

    ./verify_xlsx.py [path/to/board.xlsx]

Set FDS_SOFFICE to point at a soffice binary if it is not on PATH.
"""
from __future__ import annotations

import csv
import glob
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from build_spreadsheet import TCOL  # noqa: E402
from common import FANTASY_POS, latest_board, load_config  # noqa: E402

# 1-based column indexes on the tracker, resolved from the builder's own layout
# so inserting a column can never leave these helpers reading the wrong field.
COL = {k: ord(v) - ord("A") + 1 for k, v in TCOL.items()}

RECALC_CFG = (
    '<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>'
    '<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>'
)


def find_soffice() -> str | None:
    env = os.environ.get("FDS_SOFFICE")
    if env and Path(env).exists():
        return env
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    for pat in ("/opt/libreoffice*/program/soffice",
                "/usr/lib/libreoffice/program/soffice",
                "/Applications/LibreOffice.app/Contents/MacOS/soffice",
                r"C:\Program Files\LibreOffice\program\soffice.exe"):
        hits = glob.glob(pat)
        if hits:
            return hits[0]
    return None


def sheet_to_rows(soffice: str, xlsx: Path, sheet_index: int,
                  workdir: Path) -> list[list[str]]:
    """Recalculate the workbook and export one sheet as CSV."""
    profile = workdir / "profile"
    outdir = workdir / f"out{sheet_index}"
    outdir.mkdir(parents=True, exist_ok=True)

    # Prime the profile, then force recalculate-on-load.
    reg = profile / "user" / "registrymodifications.xcu"
    if not reg.exists():
        subprocess.run([soffice, "--headless",
                        f"-env:UserInstallation=file://{profile}", "--terminate_after_init"],
                       capture_output=True, timeout=180)
    if reg.exists():
        s = reg.read_text()
        if "OOXMLRecalcMode" not in s:
            reg.write_text(s.replace("</oor:items>", RECALC_CFG + "</oor:items>"))

    flt = ("csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,"
           f"false,false,{sheet_index}")
    subprocess.run([soffice, "--headless",
                    f"-env:UserInstallation=file://{profile}",
                    "--convert-to", flt, "--outdir", str(outdir), str(xlsx)],
                   capture_output=True, timeout=300)
    files = list(outdir.glob("*.csv"))
    if not files:
        raise RuntimeError(f"no CSV produced for sheet {sheet_index}")
    with open(files[0], newline="") as fh:
        return list(csv.reader(fh))


def _live_panel(rows: list[list[str]]) -> dict:
    """Pull the live block's key cells out of a recalculated Pick Assistant."""
    out = {}
    for r in rows:
        if len(r) < 11:
            continue
        head = r[9].strip()
        if head == "Picks made":
            out["picks_made"] = r[10].strip()
        elif head == "On the clock (pick #)":
            out["on_clock"] = r[10].strip()
        elif head == "Your NEXT turn":
            out["next_turn"] = r[10].strip()
        elif head.startswith("TAKE:") or head == "-":
            out["headline"] = head
    return out


def _mark_drafted(src: Path, dst: Path, n: int, *, mine_pos: str = None,
                  mine_n: int = 0) -> list[str]:
    """Copy the workbook and mark the n lowest-ADP players as drafted, which
    is roughly how the first n picks of a real draft go.

    `mine_pos`/`mine_n` additionally claim the best `mine_n` players at that
    position in the Mine? column, to prove the live headline respects a roster
    that is already full at a position."""
    import openpyxl
    wb = openpyxl.load_workbook(src)
    ws = wb["Live Draft Tracker"]
    rows = []
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, COL["player_name"]).value
        adp = ws.cell(r, COL["adp_10"]).value
        pos = ws.cell(r, COL["position"]).value
        if name and isinstance(adp, (int, float)):
            rows.append((adp, r, str(name), str(pos)))
    rows.sort()
    picked = rows[:n]
    for _, r, _, _ in picked:
        ws.cell(r, COL["drafted"]).value = "X"
    if mine_pos and mine_n:
        claimed = 0
        for _, r, _, pos in rows:
            if pos == mine_pos and not ws.cell(r, COL["drafted"]).value:
                ws.cell(r, COL["mine"]).value = "X"
                claimed += 1
                if claimed >= mine_n:
                    break
    wb.save(dst)
    return [nm for _, _, nm, _ in picked]


FAILS: list[str] = []
PASSES = 0


def check(name: str, cond: bool, detail: str = "") -> None:
    global PASSES
    if cond:
        PASSES += 1
        print(f"  [pass] {name}")
    else:
        FAILS.append(f"{name}: {detail}")
        print(f"  [FAIL] {name}  {detail}")


def main() -> int:
    cfg = load_config()
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if xlsx is None:
        xlsx = latest_board()
        if xlsx is None:
            print("no workbook found in output/")
            return 1

    soffice = find_soffice()
    if not soffice:
        print("LibreOffice not found - install it or set FDS_SOFFICE.")
        print("Skipping live-formula verification (this is not a pass).")
        return 2

    print(f"engine : {soffice}")
    print(f"file   : {xlsx}\n")

    import openpyxl
    names = openpyxl.load_workbook(xlsx, read_only=True).sheetnames
    idx_of = {n: i + 1 for i, n in enumerate(names)}

    with tempfile.TemporaryDirectory() as td:
        work = Path(td)
        rows = sheet_to_rows(soffice, xlsx, idx_of["Live Draft Tracker"], work)
        formula_sheets = {"Live Draft Tracker": rows}
        for extra in ("Compare", "Pick Assistant", "LiveCalc"):
            if extra in idx_of:
                formula_sheets[extra] = sheet_to_rows(
                    soffice, xlsx, idx_of[extra], work)

    print("=== formula health (all formula-bearing sheets) ===")
    for sname, srows in formula_sheets.items():
        text = "\n".join("|".join(r) for r in srows)
        check(f"{sname}: no #NAME? (missing _xlfn. prefix)", "#NAME" not in text)
        for err in ("#REF!", "#VALUE!", "#DIV/0!", "#N/A"):
            check(f"{sname}: no {err}", err not in text)

    # The Compare tab must actually resolve its INDEX/MATCH lookups.
    if "Compare" in formula_sheets:
        crows = formula_sheets["Compare"]
        pos_row = next((r for r in crows if r and r[0].strip() == "Position"), None)
        check("Compare: dropdowns resolve to real positions",
              bool(pos_row) and any(c.strip() in ("QB", "RB", "WR", "TE", "K", "DEF")
                                    for c in pos_row[1:5]),
              f"got {pos_row[1:5] if pos_row else None}")

    # The Pick Assistant must name real players and never recommend someone
    # who has essentially no chance of being on the board.
    if "Pick Assistant" in formula_sheets:
        prows = formula_sheets["Pick Assistant"]
        hdr = next((i for i, r in enumerate(prows)
                    if r and r[0].strip() == "Rank"), None)
        check("Pick Assistant: table renders", hdr is not None)
        if hdr is not None:
            body = [r for r in prows[hdr + 1:hdr + 16] if len(r) > 7 and r[1].strip()]
            check("Pick Assistant: lists players", len(body) >= 5,
                  f"only {len(body)} rows")
            bad = [r[1] for r in body
                   if r[7].strip() == "TAKE NOW"
                   and r[4].strip().rstrip("%").isdigit()
                   and int(r[4].strip().rstrip("%")) < 35]
            check("Pick Assistant: never says TAKE NOW for an unavailable player",
                  not bad, f"{bad}")

    # The probabilistic-availability engine must actually evaluate: NORMDIST
    # resolves, and survival probabilities land in [0, 1] rather than degrading
    # to all-zero (which would silently revert to a hard cutoff).
    if "LiveCalc" in formula_sheets:
        lc = formula_sheets["LiveCalc"]
        hdr = next((i for i, r in enumerate(lc)
                    if r and r[0].strip() == "pos"), None)
        check("LiveCalc: engine table renders", hdr is not None)
        if hdr is not None:
            survs = []
            for r in lc[hdr + 1:]:
                if len(r) > 5 and r[5].strip():
                    try:
                        survs.append(float(r[5]))
                    except ValueError:
                        pass
            check("LiveCalc: survival probabilities evaluate", len(survs) > 20,
                  f"only {len(survs)} numeric surv cells")
            check("LiveCalc: probabilities stay within [0, 1]",
                  all(-1e-9 <= s <= 1 + 1e-9 for s in survs),
                  f"out of range: {[s for s in survs if not -1e-9 <= s <= 1 + 1e-9][:3]}")
            check("LiveCalc: model is live, not a degenerate all-zero",
                  any(s > 0.01 for s in survs), "every survival prob is ~0")

    # --- the live suggester must actually REACT to the tracker ---
    if "Pick Assistant" in formula_sheets:
        print("\n=== live suggester reacts to draft state ===")
        base = _live_panel(formula_sheets["Pick Assistant"])
        check("live panel renders", bool(base), f"{base}")
        if base:
            print(f"  empty board -> {base.get('headline', '')[:60]}")
            check("live: agrees with an untouched board",
                  base["picks_made"] == "0",
                  f"picks_made={base['picks_made']}")

            with tempfile.TemporaryDirectory() as td:
                work = Path(td)
                marked = work / "marked.xlsx"
                names = _mark_drafted(xlsx, marked, 15)
                rows2 = sheet_to_rows(soffice, marked, idx_of["Pick Assistant"],
                                      work)
                after = _live_panel(rows2)
            print(f"  15 drafted  -> {after.get('headline', '')[:60]}")
            check("live: picks-made count follows the marks",
                  after.get("picks_made") == "15",
                  f"got {after.get('picks_made')}")
            check("live: recommendation changes once players come off",
                  after.get("headline") != base.get("headline"),
                  "headline did not move")
            check("live: no longer recommends a drafted player",
                  not any(n in after.get("headline", "") for n in names),
                  f"recommends someone already taken: {after.get('headline')}")

            # --- roster awareness: a full position must drop out ------------
            # Raw VONA loves late QBs forever; the Mine? column is what stops
            # the headline recommending a third one in a 1QB league.
            cap = int((cfg.get("output", {}).get("useful_max") or {}).get("QB", 2))
            with tempfile.TemporaryDirectory() as td:
                work = Path(td)
                claimed = work / "mine.xlsx"
                _mark_drafted(xlsx, claimed, 15, mine_pos="QB", mine_n=cap)
                prows = sheet_to_rows(soffice, claimed,
                                      idx_of["Pick Assistant"], work)
                full = _live_panel(prows)
                qb_row = next((r for r in prows
                               if len(r) > 14 and r[9].strip() == "QB"), None)
            print(f"  QB roster full -> {full.get('headline', '')[:60]}")
            def _num(x):
                try:
                    return float(str(x).strip())
                except ValueError:
                    return None
            check("live: counts your own picks from the Mine? column",
                  bool(qb_row) and _num(qb_row[14]) == float(cap),
                  f"'Yours' reads {qb_row[14] if qb_row else None}, want {cap}")
            check("live: marks a filled position ROSTER FULL",
                  bool(qb_row) and qb_row[13].strip() == "ROSTER FULL",
                  f"verdict is {qb_row[13] if qb_row else None!r}")
            check("live: headline stops recommending a filled position",
                  "(QB," not in full.get("headline", ""),
                  f"still says {full.get('headline')}")
            check("live: Mine? marks count toward picks made",
                  full.get("picks_made") == str(15 + cap),
                  f"got {full.get('picks_made')}, want {15 + cap}")

    rows = formula_sheets["Live Draft Tracker"]

    # Locate the scarcity panel and read the computed numbers.
    hdr_row = next((i for i, r in enumerate(rows)
                    if any(c.strip() == "POSITIONAL SCARCITY" for c in r)), None)
    check("scarcity panel present", hdr_row is not None)
    if hdr_row is None:
        return _report()
    col = next(i for i, c in enumerate(rows[hdr_row]) if c.strip() == "POSITIONAL SCARCITY")

    print("\n=== computed scarcity (nothing drafted) ===")
    seen = {}
    for k in range(1, len(FANTASY_POS) + 1):
        r = rows[hdr_row + k]
        pos = r[col].strip()
        left = r[col + 1].strip()
        best = r[col + 2].strip()
        in_tier = r[col + 3].strip() if len(r) > col + 3 else ""
        seen[pos] = (left, best, in_tier)
        print(f"  {pos:<4} left={left:<5} best tier={best:<4} in tier={in_tier}")

    for pos in FANTASY_POS:
        left, best, in_tier = seen.get(pos, ("", "", ""))
        check(f"{pos}: 'left' is a positive number",
              left.isdigit() and int(left) > 0, f"got {left!r}")
        check(f"{pos}: best available tier is 1 on a full board",
              best == "1", f"got {best!r}")
        check(f"{pos}: tier-1 count is positive",
              in_tier.isdigit() and int(in_tier) > 0, f"got {in_tier!r}")

    return _report()


def _report() -> int:
    print(f"\n{PASSES} passed, {len(FAILS)} failed")
    for f in FAILS:
        print(f"  - {f}")
    return 1 if FAILS else 0


if __name__ == "__main__":
    sys.exit(main())
