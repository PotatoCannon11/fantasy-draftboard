#!/usr/bin/env python
"""Mirror a live Sleeper draft into the workbook, and call your picks.

Sleeper's mock drafts let you draft against their bots. Their API is read-only
and unauthenticated, so we cannot make picks for you - but we can watch one and
keep the board in step with it, which is the part that matters:

    1. Start a mock at https://sleeper.com/mockdraft
    2. Copy the draft id out of the URL (.../draft/nfl/<draft_id>)
    3. ./sleeper_sync.py <draft_id> --slot 5

Every pick is mapped onto a board player and marked in a COPY of the workbook -
yours in `Mine?`, everyone else's in `Drafted?` - and when you are on the clock
the Pick Assistant is recalculated and its call printed to the terminal.

The copy is written to output/live_draft_<date>.xlsx so your master file is
never touched. LibreOffice does not reload a file changed underneath it: keep
watching the terminal during the draft, and use File > Reload if you want to see
the sheet itself.

    ./sleeper_sync.py <draft_id> --slot 5           # follow a live draft
    ./sleeper_sync.py <draft_id> --once             # one pass and exit
    ./sleeper_sync.py --replay picks.json --slot 5  # offline, from a saved payload
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import date
from pathlib import Path

import pandas as pd
import requests

sys.path.insert(0, str(Path(__file__).resolve().parent))

from common import DATA_PROC, OUTPUT, latest_board, load_config, norm_name  # noqa: E402
from idmap import IdResolver  # noqa: E402
from verify_xlsx import COL, find_soffice, sheet_to_rows  # noqa: E402

DRAFT_URL = "https://api.sleeper.app/v1/draft/{draft_id}"
PICKS_URL = "https://api.sleeper.app/v1/draft/{draft_id}/picks"
SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0 (fantasy-draft-system)"})


def _get(url: str, timeout: int = 20):
    last = None
    for attempt in range(3):
        try:
            r = SESSION.get(url, timeout=timeout)
            r.raise_for_status()
            return r.json()
        except Exception as exc:  # noqa: BLE001
            last = exc
            time.sleep(1.2 * (attempt + 1))
    raise RuntimeError(f"failed to fetch {url}: {last}")


def parse_draft_id(s: str) -> str:
    """Accept a bare id or any Sleeper URL containing one."""
    s = s.strip()
    m = re.search(r"(\d{6,})", s)
    if not m:
        raise SystemExit(f"could not find a draft id in {s!r}")
    return m.group(1)


def _norm_sid(x) -> str | None:
    """resolver.by_sleeper keys are float-stringified ('96.0'); the draft feed
    returns plain ids ('11571'). Normalize both sides to integer strings."""
    try:
        return str(int(float(x)))
    except (TypeError, ValueError):
        return str(x) if x is not None else None


def fetch_draft(draft_id: str) -> dict:
    return _get(DRAFT_URL.format(draft_id=draft_id))


def fetch_picks(draft_id: str) -> list[dict]:
    picks = _get(PICKS_URL.format(draft_id=draft_id)) or []
    return sorted(picks, key=lambda p: int(p.get("pick_no") or 0))


class BoardMap:
    """Maps a Sleeper pick onto a board player.

    `xlsx` is optional: the spreadsheet path additionally builds tracker row
    numbers (used by sleeper_sync's mirror). The dashboard only needs names, so
    it constructs this without a workbook."""

    def __init__(self, xlsx: Path | None = None, board: pd.DataFrame = None):
        if board is None:
            board = pd.read_parquet(DATA_PROC / "draft_board.parquet")
        resolver = IdResolver()

        self.uid_to_name = dict(zip(board.player_uid, board.player_name))
        self.by_norm = {norm_name(n): n for n in board.player_name}
        self.sid_to_uid: dict[str, str] = {}
        for k, uid in resolver.by_sleeper.items():
            nk = _norm_sid(k)
            if nk:
                self.sid_to_uid.setdefault(nk, uid)

        # Team defenses: Sleeper's player_id IS the team code ("SF"), and the
        # board names them "SF DEF", so they never resolve through the id map.
        self.def_names = {str(n).split()[0].upper(): n
                          for n in board.loc[board.position == "DEF",
                                             "player_name"]}

        self.row_of: dict[str, int] = {}
        if xlsx is not None:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx, read_only=True)
            ws = wb["Live Draft Tracker"]
            for r, row in enumerate(
                    ws.iter_rows(min_row=5, max_col=COL["player_name"],
                                 values_only=True), start=5):
                nm = row[COL["player_name"] - 1]
                if nm:
                    self.row_of.setdefault(str(nm).strip(), r)
            wb.close()

    def resolve_name(self, pick: dict) -> str | None:
        """Board player name for a Sleeper pick, or None if unmatched."""
        meta = pick.get("metadata") or {}
        pid = str(pick.get("player_id") or "")

        # 1. Team defense, by team code - Sleeper's player_id IS "SF".
        if (meta.get("position") or "").upper() in ("DEF", "DST"):
            nm = self.def_names.get(pid.upper()) or \
                self.def_names.get((meta.get("team") or "").upper())
            if nm:
                return nm

        # 2. Canonical id map.
        uid = self.sid_to_uid.get(_norm_sid(pid))
        nm = self.uid_to_name.get(uid) if uid else None
        if nm is not None:
            return nm

        # 3. Fall back to the name Sleeper sent with the pick.
        full = f"{meta.get('first_name', '')} {meta.get('last_name', '')}".strip()
        if full:
            return self.by_norm.get(norm_name(full))
        return None

    def describe(self, pick: dict) -> str:
        """Best human label for a pick, matched or not."""
        meta = pick.get("metadata") or {}
        full = f"{meta.get('first_name', '')} {meta.get('last_name', '')}".strip()
        return self.resolve_name(pick) or full or str(pick.get("player_id"))

    def resolve(self, pick: dict) -> tuple[str | None, int | None]:
        """(board player name, tracker row). Row is None without an xlsx."""
        nm = self.resolve_name(pick)
        return (nm or self.describe(pick)), (self.row_of.get(nm) if nm else None)


def apply_picks(src: Path, dst: Path, picks: list[dict], bmap: BoardMap,
                my_slot: int) -> tuple[int, list[str]]:
    """Write every pick into a copy of the workbook. Yours land in `Mine?`,
    the room's in `Drafted?` - the same one-mark-each convention the sheet
    expects. Returns (marks written, unmatched descriptions)."""
    import openpyxl
    wb = openpyxl.load_workbook(src)
    ws = wb["Live Draft Tracker"]
    marked, missing = 0, []
    for p in picks:
        name, row = bmap.resolve(p)
        if row is None:
            missing.append(f"pick {p.get('pick_no')}: {name}")
            continue
        col = COL["mine"] if int(p.get("draft_slot") or -1) == my_slot \
            else COL["drafted"]
        ws.cell(row, col).value = "X"
        marked += 1
    wb.save(dst)
    wb.close()
    return marked, missing


def read_live_panel(soffice: str, xlsx: Path, workdir: Path) -> dict:
    import openpyxl
    names = openpyxl.load_workbook(xlsx, read_only=True).sheetnames
    idx = {n: i + 1 for i, n in enumerate(names)}
    rows = sheet_to_rows(soffice, xlsx, idx["Pick Assistant"], workdir)
    out = {"positions": []}
    for r in rows:
        if len(r) < 11:
            continue
        head = r[9].strip()
        if head == "Picks made":
            out["picks_made"] = r[10].strip()
        elif head == "Your NEXT turn":
            out["next_turn"] = r[10].strip()
        elif head.startswith("TAKE:") or head.startswith("ROSTER FULL"):
            out["headline"] = head
        elif head in ("QB", "RB", "WR", "TE", "K", "DEF") and len(r) > 14:
            out["positions"].append(
                (head, r[10].strip(), r[12].strip(), r[13].strip(),
                 r[14].strip()))
    return out


def snake_slot(pick_no: int, teams: int) -> int:
    """Which draft slot owns an overall pick number, snake order."""
    rnd, i = divmod(pick_no - 1, teams)
    return i + 1 if rnd % 2 == 0 else teams - i


def main(argv=None) -> int:
    cfg = load_config()
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("draft", nargs="?", help="Sleeper draft id or URL")
    ap.add_argument("--slot", type=int, default=cfg["simulation"]["draft_slot"],
                    help="your seat in the mock (1 = first overall)")
    ap.add_argument("--interval", type=float, default=5.0)
    ap.add_argument("--once", action="store_true", help="one pass, then exit")
    ap.add_argument("--replay", type=Path,
                    help="read a saved picks JSON instead of the network")
    ap.add_argument("--xlsx", type=Path, help="master workbook to copy from")
    ap.add_argument("--out", type=Path, help="where to write the live copy")
    args = ap.parse_args(argv)

    if not args.draft and not args.replay:
        ap.error("give a draft id/URL, or --replay a saved payload")

    src = args.xlsx
    if src is None:
        src = latest_board()
        if src is None:
            print("no workbook in output/ - run the pipeline first")
            return 1
    dst = args.out or (OUTPUT / f"live_draft_{date.today().isoformat()}.xlsx")
    if Path(src).resolve() == Path(dst).resolve():
        print("refusing to use the live mirror as its own master - pass --xlsx")
        return 1

    soffice = find_soffice()
    teams = cfg["league"]["primary_team_count"]
    rounds = cfg["simulation"]["rounds"]

    draft_id = None
    if args.draft:
        draft_id = parse_draft_id(args.draft)
        info = fetch_draft(draft_id)
        st = info.get("settings") or {}
        teams = int(st.get("teams") or teams)
        rounds = int(st.get("rounds") or rounds)
        print(f"draft   : {draft_id}  ({info.get('type')}, status "
              f"{info.get('status')})")
        if teams != cfg["league"]["primary_team_count"]:
            print(f"  !! this mock has {teams} teams but the board is tiered for "
                  f"{cfg['league']['primary_team_count']} - VBD and ADP columns "
                  f"are for the configured size, so read them with that in mind")

    print(f"master  : {src}")
    print(f"live    : {dst}")
    print(f"seat    : slot {args.slot} of {teams}, {rounds} rounds")
    if not soffice:
        print("  !! LibreOffice not found - marks will still be written, but the "
              "Pick Assistant cannot be recalculated. Set FDS_SOFFICE.")
    print()

    bmap = BoardMap(xlsx=src)
    import tempfile
    seen = 0
    with tempfile.TemporaryDirectory() as td:
        work = Path(td)
        while True:
            if args.replay:
                picks = sorted(json.loads(args.replay.read_text()),
                               key=lambda p: int(p.get("pick_no") or 0))
            else:
                picks = fetch_picks(draft_id)

            if len(picks) != seen:
                for p in picks[seen:]:
                    name, row = bmap.resolve(p)
                    slot = int(p.get("draft_slot") or -1)
                    who = "YOU " if slot == args.slot else "    "
                    flag = "" if row else "   <-- not on board"
                    print(f"  {who}#{p.get('pick_no'):>3} R{p.get('round')} "
                          f"slot {slot:<2} {name}{flag}")
                seen = len(picks)

                marked, missing = apply_picks(src, dst, picks, bmap, args.slot)
                if missing:
                    print(f"  ({len(missing)} pick(s) could not be matched to a "
                          f"board row: {'; '.join(missing[:3])})")

                nxt = len(picks) + 1
                on_clock = snake_slot(nxt, teams)
                if soffice and on_clock == args.slot:
                    panel = read_live_panel(soffice, dst, work)
                    print("\n  " + "=" * 58)
                    print(f"  YOU ARE ON THE CLOCK - pick #{nxt} (round "
                          f"{(nxt - 1) // teams + 1})")
                    print(f"  {panel.get('headline', '(no call)')}")
                    for pos, best, verdict, yours, _extra in [
                            (p[0], p[1], p[3], p[4], None)
                            for p in panel["positions"]]:
                        print(f"    {pos:<4} {best:<24} {verdict:<12} "
                              f"yours={yours}")
                    print("  " + "=" * 58 + "\n")

            done = len(picks) >= teams * rounds
            if args.once or args.replay or done:
                if done:
                    print("\ndraft complete.")
                break
            time.sleep(args.interval)

    print(f"\nmarks written to {dst}")
    print("LibreOffice will not notice the change on its own - File > Reload.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
