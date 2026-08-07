#!/usr/bin/env python
"""Run mock drafts against the REAL workbook and score the resulting rosters.

This is a draft-mechanics test, not a projection-accuracy test. Read this before
trusting the numbers it prints:

  * Opponents draft from live FFC ADP with its per-player stdev, using the same
    roster-aware model `fantasydraft/simulate.py` uses (starter nudges, positional caps,
    K/DEF held to the endgame). That part is a fair imitation of a real room.

  * YOUR picks in the `workbook` strategy come from the actual xlsx - each round
    the drafted players are marked in the Live Draft Tracker, LibreOffice
    recalculates, and we read whatever the Pick Assistant's LIVE headline says
    to take. So this exercises the artifact you use on draft day, formulas and
    all, not a Python re-implementation of it.

  * Rosters are scored with the board's OWN `final_projection`. The board is
    therefore graded on the objective it optimises, which flatters it. What this
    can legitimately show is whether the draft-decision layer converts a fixed
    set of beliefs into more points than naive strategies do, and whether the
    live formulas hold up over 15 rounds of real draft state. It cannot tell you
    whether the projections are right - only the 2026 season can do that.

  * `--score-ly` additionally scores each roster on 2025 actual PPR points as a
    cross-check that does not come from the model's output. It is a weak proxy
    (it ignores every 2026 change and has no value for rookies), so treat a
    disagreement between the two scorings as a signal about where the edge comes
    from, not as a verdict.

Usage:
    FDS_SOFFICE=/opt/libreoffice26.2/program/soffice ./mock_draft.py --drafts 10
"""
from __future__ import annotations

import argparse
import re
import sys
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))

from common import DATA_PROC, FANTASY_POS, latest_board, load_config  # noqa: E402
from simulate import my_picks, team_of_pick  # noqa: E402
from verify_xlsx import COL, find_soffice, sheet_to_rows  # noqa: E402

TAKE_RE = re.compile(r"TAKE:\s*(.+?)\s*\(")


# ---------------------------------------------------------------------------
# Draft state
# ---------------------------------------------------------------------------
class Draft:
    """One mock draft. Opponents are ADP bots; seat `slot` is driven by a
    strategy callback so different strategies face an identical room."""

    def __init__(self, board: pd.DataFrame, cfg: dict, *, n_teams: int,
                 slot: int, rounds: int, seed: int):
        self.board = board.reset_index(drop=True)
        self.cfg = cfg
        self.n_teams, self.slot, self.rounds = n_teams, slot, rounds
        self.opp = {"start_bonus": 10, "flex_bonus": 5, "kdef_last_rounds": 3,
                    "caps": {"QB": 3, "RB": 7, "WR": 8, "TE": 3, "K": 2, "DEF": 2}}
        self.opp.update(cfg.get("simulation", {}).get("opponents") or {})
        self.roster_cfg = cfg["league"]["roster"]
        self.flex_elig = cfg["league"]["flex_eligible"]

        adp = pd.to_numeric(board[f"adp_{n_teams}"], errors="coerce")
        sd = pd.to_numeric(board.get(f"adp_sd_{n_teams}"), errors="coerce")
        total = n_teams * rounds
        self.adp = adp.fillna(total + 25).to_numpy(float)
        sd = sd.to_numpy(float)
        self.sd = np.where(np.isnan(sd) | (sd <= 0),
                           np.maximum(4.0, self.adp * 0.28), sd)

        # Common random numbers: the room's preference order is identical across
        # strategies, so any difference in outcome comes from YOUR picks, not
        # from a luckier draw.
        rng = np.random.default_rng(seed)
        self.draws = self.adp + rng.standard_normal(len(self.adp)) * self.sd

        self.pos = board["position"].to_numpy()
        self.names = board["player_name"].to_numpy()
        self.avail = np.ones(len(board), dtype=bool)
        self.counts = np.zeros((n_teams, len(FANTASY_POS)), dtype=int)
        self.my_seat = slot - 1
        self.my_picks = set(my_picks(n_teams, slot, rounds))
        self.teams = team_of_pick(n_teams, rounds)
        self.taken_order: list[int] = []      # board indices, in pick order
        self.my_roster: list[int] = []

    # -- roster helpers ----------------------------------------------------
    def _pos_i(self, i: int) -> int:
        p = self.pos[i]
        return FANTASY_POS.index(p) if p in FANTASY_POS else 0

    def legal(self, team: int, rnd: int) -> np.ndarray:
        """Players this team may still take: available, under the positional
        cap, and - in the last rounds - restricted to mandatory slots it is
        still missing, which is what stops a greedy bot ending with no kicker."""
        ok = self.avail.copy()
        caps = self.opp["caps"]
        cnt = self.counts[team]
        for pi, p in enumerate(FANTASY_POS):
            if cnt[pi] >= caps.get(p, 99):
                ok &= self.pos != p
        missing = [p for p in FANTASY_POS
                   if self.counts[team][FANTASY_POS.index(p)] < self.roster_cfg.get(p, 0)]
        left = self.rounds - rnd + 1
        if missing and left <= len(missing):
            ok &= np.isin(self.pos, missing)
        return ok

    def commit(self, team: int, idx: int) -> None:
        self.avail[idx] = False
        self.counts[team, self._pos_i(idx)] += 1
        self.taken_order.append(idx)
        if team == self.my_seat:
            self.my_roster.append(idx)

    def opponent_pick(self, team: int, rnd: int) -> int:
        """Mirror of simulate._roster_aware for a single draft."""
        need = np.zeros(len(FANTASY_POS))
        cnt = self.counts[team]
        for pi, p in enumerate(FANTASY_POS):
            if cnt[pi] < self.roster_cfg.get(p, 0):
                need[pi] = self.opp["start_bonus"]
        flex_have = sum(cnt[FANTASY_POS.index(p)] for p in self.flex_elig)
        flex_total = sum(self.roster_cfg.get(p, 0) for p in self.flex_elig) \
            + self.roster_cfg.get("FLEX", 0)
        if flex_have < flex_total:
            for p in self.flex_elig:
                need[FANTASY_POS.index(p)] += self.opp["flex_bonus"]
        if rnd < self.rounds - self.opp["kdef_last_rounds"] + 1:
            for p in ("K", "DEF"):
                if p in FANTASY_POS:
                    need[FANTASY_POS.index(p)] = 0.0

        eff = self.draws - np.array([need[self._pos_i(i)]
                                     for i in range(len(self.draws))])
        ok = self.legal(team, rnd)
        eff = np.where(ok, eff, np.inf)
        if not np.isfinite(eff).any():                     # exhausted: anything
            eff = np.where(self.avail, self.draws, np.inf)
        return int(eff.argmin())


# ---------------------------------------------------------------------------
# Strategies for YOUR seat
# ---------------------------------------------------------------------------
def strat_adp(d: Draft, rnd: int, _ctx) -> int:
    """What the room does: best-available by ADP."""
    ok = d.legal(d.my_seat, rnd)
    return int(np.where(ok, d.adp, np.inf).argmin())


def strat_vbd(d: Draft, rnd: int, _ctx) -> int:
    """Best value-over-replacement available, ignoring when he'd be gone."""
    v = pd.to_numeric(d.board["vbd_score"], errors="coerce").fillna(-1e9).to_numpy()
    ok = d.legal(d.my_seat, rnd)
    return int(np.where(ok, v, -np.inf).argmax())


def strat_value(d: Draft, rnd: int, _ctx) -> int:
    """Biggest gap between the board and the market."""
    v = pd.to_numeric(d.board["value_vs_adp"], errors="coerce").fillna(-1e9).to_numpy()
    ok = d.legal(d.my_seat, rnd)
    return int(np.where(ok, v, -np.inf).argmax())


class WorkbookStrategy:
    """Drives the real xlsx: mark drafted players, recalculate, do what the
    Pick Assistant's LIVE headline says."""

    def __init__(self, xlsx: Path, soffice: str, workdir: Path, verbose=False):
        import openpyxl
        self.xlsx, self.soffice, self.workdir = xlsx, soffice, workdir
        self.verbose = verbose
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        self.sheet_idx = {n: i + 1 for i, n in enumerate(wb.sheetnames)}
        ws = wb["Live Draft Tracker"]
        # Resolve columns from the builder's layout, never by hardcoded index.
        self.c_name = COL["player_name"]
        self.c_drafted, self.c_mine = COL["drafted"], COL["mine"]
        self.row_of: dict[str, int] = {}
        for r, row in enumerate(ws.iter_rows(min_row=5, max_col=self.c_name,
                                             values_only=True), start=5):
            if row[self.c_name - 1]:
                self.row_of.setdefault(str(row[self.c_name - 1]).strip(), r)
        wb.close()
        self.fallbacks = 0
        self.canary_fails = 0

    def __call__(self, d: Draft, rnd: int, _ctx) -> int:
        import openpyxl
        marked = [d.names[i] for i in d.taken_order]
        mine = {d.names[i] for i in d.my_roster}
        wb = openpyxl.load_workbook(self.xlsx)
        ws = wb["Live Draft Tracker"]
        # Exactly how you would run it live: your picks go in Mine?, everyone
        # else's in Drafted?, never both.
        for nm in marked:
            r = self.row_of.get(str(nm).strip())
            if r:
                col = self.c_mine if nm in mine else self.c_drafted
                ws.cell(r, col).value = "X"
        tmp = self.workdir / "live.xlsx"
        wb.save(tmp)
        wb.close()

        rows = sheet_to_rows(self.soffice, tmp, self.sheet_idx["Pick Assistant"],
                             self.workdir)
        headline, picks_made = "", None
        for r in rows:
            if len(r) < 11:
                continue
            if r[9].strip() == "Picks made":
                picks_made = r[10].strip()
            if r[9].strip().startswith("TAKE:"):
                headline = r[9].strip()

        # Canary: if the sheet did not actually recalculate, `picks_made` will
        # not track our marks and every recommendation below is meaningless.
        if picks_made != str(len(marked)):
            self.canary_fails += 1

        m = TAKE_RE.search(headline)
        if m:
            name = m.group(1).strip()
            hit = np.where((d.names == name) & d.avail)[0]
            if len(hit):
                idx = int(hit[0])
                if d.legal(d.my_seat, rnd)[idx]:
                    if self.verbose:
                        print(f"      R{rnd:<2} sheet says: {headline[:58]}")
                    return idx
        self.fallbacks += 1
        if self.verbose:
            print(f"      R{rnd:<2} sheet gave no usable pick "
                  f"({headline[:40]!r}) -> VBD fallback")
        return strat_vbd(d, rnd, None)


# ---------------------------------------------------------------------------
def run_draft(board, cfg, strategy, *, n_teams, slot, rounds, seed) -> Draft:
    d = Draft(board, cfg, n_teams=n_teams, slot=slot, rounds=rounds, seed=seed)
    for pick in range(1, n_teams * rounds + 1):
        rnd = (pick - 1) // n_teams + 1
        team = int(d.teams[pick - 1])
        if pick in d.my_picks:
            idx = strategy(d, rnd, None)
        else:
            idx = d.opponent_pick(team, rnd)
        d.commit(team, idx)
    return d


def lineup_points(board: pd.DataFrame, roster: list[int], cfg: dict,
                  col: str) -> tuple[float, bool]:
    """Best legal starting lineup from a roster. Returns (points, is_legal)."""
    pts = pd.to_numeric(board[col], errors="coerce").fillna(0.0).to_numpy()
    pos = board["position"].to_numpy()
    r = cfg["league"]["roster"]
    pool = {p: sorted((pts[i] for i in roster if pos[i] == p), reverse=True)
            for p in FANTASY_POS}
    total, legal = 0.0, True
    used = {p: 0 for p in FANTASY_POS}
    for p in FANTASY_POS:
        want = r.get(p, 0)
        have = pool[p][:want]
        if len(have) < want:
            legal = False
        total += sum(have)
        used[p] = len(have)
    for _ in range(r.get("FLEX", 0)):  # noqa: B007
        best, bp = None, None
        for p in cfg["league"]["flex_eligible"]:
            rem = pool[p][used[p]:]
            if rem and (best is None or rem[0] > best):
                best, bp = rem[0], p
        if best is None:
            legal = False
        else:
            total += best
            used[bp] += 1
    return total, legal


# In a 1QB league a second QB is a sane bench stash and a third is dead weight;
# same for K/DEF past one and TE past two. RB/WR depth always has a use, so it is
# never counted as waste.
USEFUL_MAX = {"QB": 2, "TE": 2, "K": 1, "DEF": 1}


def wasted_picks(board: pd.DataFrame, roster: list[int]) -> int:
    """Picks spent on a position the roster can no longer use. This is the cost
    of a recommender that cannot see what you have already drafted."""
    pos = board["position"].to_numpy()
    have: dict[str, int] = {}
    waste = 0
    for i in roster:
        p = pos[i]
        have[p] = have.get(p, 0) + 1
        if p in USEFUL_MAX and have[p] > USEFUL_MAX[p]:
            waste += 1
    return waste


def main(argv=None) -> int:
    cfg = load_config()
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--drafts", type=int, default=10)
    ap.add_argument("--slot", type=int, default=cfg["simulation"]["draft_slot"])
    ap.add_argument("--teams", type=int, default=cfg["league"]["primary_team_count"])
    ap.add_argument("--rounds", type=int, default=cfg["simulation"]["rounds"])
    ap.add_argument("--xlsx", type=Path, default=None)
    ap.add_argument("--score-ly", action="store_true",
                    help="also score rosters on 2025 actual PPR points")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args(argv)

    board = pd.read_parquet(DATA_PROC / "draft_board.parquet")
    xlsx = args.xlsx

    if xlsx is None:
        xlsx = latest_board()
        if xlsx is None:
            print("no workbook in output/ - run the pipeline first")
            return 1
    soffice = find_soffice()
    if not soffice:
        print("LibreOffice not found - set FDS_SOFFICE. Cannot drive the workbook.")
        return 2

    print(f"engine  : {soffice}")
    print(f"workbook: {xlsx}")
    print(f"room    : {args.teams} teams, slot {args.slot}, {args.rounds} rounds, "
          f"{args.drafts} drafts (common random numbers)")

    with tempfile.TemporaryDirectory() as td:
        work = Path(td)
        wbs = WorkbookStrategy(xlsx, soffice, work, verbose=args.verbose)

        # The sheet IS the universe: a player the workbook does not list cannot
        # be marked, so letting opponents draft one would desync the tracker's
        # pick count from reality and quietly invalidate every LIVE number.
        # Anyone dropped here is a player you could not have tracked on draft
        # day either - which is a finding about the workbook, not the harness.
        on_sheet = board["player_name"].astype(str).str.strip().isin(wbs.row_of)
        drafted_range = pd.to_numeric(board["adp_10"], errors="coerce") <= \
            args.teams * args.rounds
        lost = board[~on_sheet & drafted_range.fillna(False)]
        if len(lost):
            print(f"\n  !! {len(lost)} player(s) inside the drafted range are "
                  f"MISSING from the workbook and cannot be tracked:")
            for _, r in lost.iterrows():
                print(f"       {r.player_name} ({r.position}, ADP "
                      f"{r['adp_10']}, VBD rank {int(r.overall_rank)})")
        board = board[on_sheet].reset_index(drop=True)
        print(f"  universe: {len(board)} players listed on the sheet\n")
        strategies = {
            "workbook (LIVE)": wbs,
            "best VBD": strat_vbd,
            "best value vs ADP": strat_value,
            "ADP (the room)": strat_adp,
        }
        results: dict[str, list] = {k: [] for k in strategies}
        rosters: dict[str, Draft] = {}

        for n in range(args.drafts):
            seed = 1000 + n
            if args.verbose:
                print(f"  -- draft {n + 1} (seed {seed}) --")
            for label, fn in strategies.items():
                d = run_draft(board, cfg, fn, n_teams=args.teams, slot=args.slot,
                              rounds=args.rounds, seed=seed)
                proj, legal = lineup_points(board, d.my_roster, cfg,
                                            "final_projection")
                row = {"proj": proj, "legal": legal,
                       "waste": wasted_picks(board, d.my_roster)}
                if args.score_ly:
                    ly, _ = lineup_points(board, d.my_roster, cfg, "ppr_points_ly")
                    row["ly"] = ly
                results[label].append(row)
                if n == 0:
                    rosters[label] = d
            print(f"  draft {n + 1}/{args.drafts} done", flush=True)

    print("\n=== starting-lineup points, board's own projections ===")
    print(f"  {'strategy':<20} {'mean':>8} {'sd':>7} {'min':>8} {'max':>8} "
          f"{'vs ADP':>8} {'illegal':>8} {'wasted':>8}")
    base = float(np.mean([r["proj"] for r in results["ADP (the room)"]]))
    for label, rs in results.items():
        v = np.array([r["proj"] for r in rs])
        bad = sum(1 for r in rs if not r["legal"])
        w = float(np.mean([r["waste"] for r in rs]))
        print(f"  {label:<20} {v.mean():>8.1f} {v.std():>7.1f} {v.min():>8.1f} "
              f"{v.max():>8.1f} {v.mean() - base:>+8.1f} {bad:>8} {w:>8.1f}")
    print("  'wasted' = mean picks spent on a position the roster can no longer "
          "use (QB3+, TE3+, K2+, DEF2+)")

    if args.score_ly:
        print("\n=== same rosters, scored on 2025 ACTUAL PPR points (cross-check) ===")
        print("  (independent of the model's output, but blind to every 2026 change)")
        base_ly = float(np.mean([r["ly"] for r in results["ADP (the room)"]]))
        for label, rs in results.items():
            v = np.array([r["ly"] for r in rs])
            print(f"  {label:<20} {v.mean():>8.1f} {v.std():>7.1f} "
                  f"{v.mean() - base_ly:>+8.1f}")

    print(f"\nworkbook health: {wbs.fallbacks} pick(s) where the LIVE headline "
          f"was unusable, {wbs.canary_fails} recalc canary failure(s)")
    if wbs.canary_fails:
        print("  !! the sheet did not track the marks - results above are void")

    d0 = rosters["workbook (LIVE)"]
    print(f"\n=== roster the workbook built in draft 1 (slot {args.slot}) ===")
    for k, i in enumerate(d0.my_roster, 1):
        b = board.iloc[i]
        print(f"  R{k:<3} {b.player_name:<24} {b.position:<4} "
              f"proj={b.final_projection:6.1f}  ADP={b[f'adp_{args.teams}']}")
    return 1 if wbs.canary_fails else 0


if __name__ == "__main__":
    sys.exit(main())
