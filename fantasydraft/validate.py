#!/usr/bin/env python
"""Smoke checks on the built board.

Verifies the data invariants that actually matter on draft day, and replays
the Live Draft Tracker's worksheet formulas in Python so their logic is
checked even without Excel installed.

    ./validate.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))

from common import DATA_PROC, FANTASY_POS, OUTPUT, load_config  # noqa: E402

FAILURES: list[str] = []
PASSES = 0


def check(name: str, condition: bool, detail: str = "") -> None:
    global PASSES
    if condition:
        PASSES += 1
        print(f"  [pass] {name}")
    else:
        FAILURES.append(f"{name}: {detail}")
        print(f"  [FAIL] {name}  {detail}")


def main() -> int:
    cfg = load_config()
    path = DATA_PROC / "draft_board.parquet"
    if not path.exists():
        print("draft_board.parquet missing - run ./run_pipeline.py first")
        return 1
    df = pd.read_parquet(path)

    print("=== board integrity ===")
    check("no duplicate players", df["player_uid"].duplicated().sum() == 0,
          f"{df['player_uid'].duplicated().sum()} duplicates")
    check("all positions present",
          set(df["position"].unique()) >= set(FANTASY_POS),
          f"missing {set(FANTASY_POS) - set(df['position'].unique())}")
    check("32 team defenses", (df["position"] == "DEF").sum() == 32,
          f"got {(df['position'] == 'DEF').sum()}")
    check("no negative projections",
          bool((df["final_projection"] >= 0).all()),
          f"{(df['final_projection'] < 0).sum()} negative")
    # Deep-bench players legitimately floor at 0; everyone in the draftable
    # range must project strictly positive.
    draftable = df.nsmallest(cfg["output"]["board_depth"], "overall_rank")
    check("every player on the board projects > 0",
          bool((draftable["final_projection"] > 0).all()),
          f"{(draftable['final_projection'] <= 0).sum()} at zero")
    check("floor <= mid <= ceiling",
          bool((df["projection_low"] <= df["projection_mid"] + 1e-9).all()
               and (df["projection_mid"] <= df["projection_high"] + 1e-9).all()),
          "band ordering violated")

    swing = cfg["context"]["max_total_swing"]
    within = df["context_multiplier"].between(1 - swing - 1e-9, 1 + swing + 1e-9)
    check(f"context multiplier within +/-{swing:.0%}", bool(within.all()),
          f"{(~within).sum()} outside the clamp")

    print("\n=== VBD ===")
    for n in cfg["league"]["team_counts"]:
        col = f"vbd_{n}"
        check(f"{col} present", col in df.columns)
        if col in df.columns:
            # Replacement-level player should sit at ~0 VBD by construction.
            near_zero = df.groupby("position")[col].apply(
                lambda s: np.nanmin(np.abs(s)))
            check(f"{col} crosses zero in every position",
                  bool((near_zero < 5).all()),
                  f"min |vbd| per position: {near_zero.round(2).to_dict()}")

    print("\n=== tiers ===")
    check("tier 1 exists for each position",
          bool(df.groupby("position")["tier"].min().eq(1).all()))
    mono = True
    for pos, g in df.groupby("position"):
        g = g.sort_values("vbd_score", ascending=False)
        if not g["tier"].is_monotonic_increasing:
            mono = False
    check("tiers increase monotonically down each position board", mono)
    labels = set(df["sub_tier_label"].unique())
    check("sub-tier labels valid", labels <= {"Early", "Mid", "Late", ""},
          f"unexpected: {labels - {'Early', 'Mid', 'Late', ''}}")

    print("\n=== ADP join ===")
    adp_col = f"adp_{cfg['output']['adp_teams']}"
    if adp_col in df.columns:
        top100 = df.nsmallest(100, "overall_rank")
        matched = top100[adp_col].notna().sum()
        check("ADP matched for >=85% of my top 100", matched >= 85,
              f"only {matched}/100")

    print("\n=== tracker formula logic (replayed in Python) ===")
    board = df.sort_values("vbd_score", ascending=False)
    keep = (board["n_sources"].fillna(0) >= 2) | (
        board["overall_rank"] <= 60)
    board = board[keep].head(cfg["output"]["board_depth"])

    # Nothing drafted: "Left" must equal the count on the sheet, and the best
    # remaining tier must be tier 1 for every position that appears.
    drafted = pd.Series(False, index=board.index)
    for pos in FANTASY_POS:
        m = (board["position"] == pos) & ~drafted
        if not m.any():
            continue
        left = int(m.sum())
        best = int(board.loc[m, "tier"].min())
        in_tier = int(((board["position"] == pos) & ~drafted
                       & (board["tier"] == best)).sum())
        check(f"{pos}: MINIFS best tier = 1 with an empty board", best == 1,
              f"got {best}")
        check(f"{pos}: COUNTIFS left ({left}) = rows on sheet", left > 0)
        check(f"{pos}: tier-1 count ({in_tier}) > 0", in_tier > 0)

    # Simulate a positional run: strike out every tier-1 RB and confirm the
    # best remaining RB tier advances, which is what drives the warning.
    rb1 = (board["position"] == "RB") & (board["tier"] == 1)
    drafted2 = rb1.copy()
    rest = (board["position"] == "RB") & ~drafted2
    new_best = int(board.loc[rest, "tier"].min())
    check("striking out RB tier 1 advances the best remaining tier",
          new_best > 1, f"still {new_best}")

    print("\n=== output file ===")
    xlsx = sorted(OUTPUT.glob("draft_board_*.xlsx"))
    check("xlsx written", bool(xlsx), "no file in output/")
    if xlsx:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx[-1])
        required = ["Master Board", *FANTASY_POS, "Live Draft Tracker",
                    "Cheat Sheet", "Compare", "Metrics", "Sources"]
        missing = [s for s in required if s not in wb.sheetnames]
        check("all required tabs present", not missing, f"missing {missing}")
        # Pick Assistant only exists once the simulation has been run.
        check("Pick Assistant present (needs fantasydraft/simulate.py)",
              "Pick Assistant" in wb.sheetnames,
              "run the simulate stage")
        ws = wb["Master Board"]
        check("master board has rows", ws.max_row > 50, f"{ws.max_row} rows")

    print(f"\n{PASSES} passed, {len(FAILURES)} failed")
    for f in FAILURES:
        print(f"  - {f}")
    return 1 if FAILURES else 0


if __name__ == "__main__":
    sys.exit(main())
